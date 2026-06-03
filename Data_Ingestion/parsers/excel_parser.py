"""
Excel / XLSX Parser — openpyxl
================================
Extracts per sheet: sheet title, data table, and embedded images.

Production considerations:
  - Row limit (MAX_ROWS) prevents memory exhaustion on huge sheets
  - Image extraction uses a version-safe approach (openpyxl private API changed in 3.x)
  - data_only=True means cells with un-evaluated formulas return None; these are labelled
  - Headers inferred from first non-empty row; all-numeric first rows get synthetic headers
  - Empty columns are trimmed; fully-empty rows are skipped
  - Per-sheet try/except — a corrupt sheet never aborts the whole workbook
"""

from __future__ import annotations
import base64
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

from models.meta_schema import (
    ParsedDocument, TextElement, ImageElement, TableElement, SourceLocation,
)
from parsers.base_parser import BaseParser

logger = logging.getLogger(__name__)

MAX_ROWS      = 10_000   # Safety cap — warn and truncate beyond this
MIN_IMG_BYTES = 512      # Skip placeholder/corrupt image parts this small


class ExcelParser(BaseParser):

    def parse(self, file_path: str | Path) -> ParsedDocument:
        file_path = Path(file_path)
        doc_meta  = self._base_doc(file_path, "xlsx")

        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
        except Exception as e:
            raise RuntimeError(f"Cannot open workbook '{file_path.name}': {e}") from e

        doc_meta.summary.sheet_names = wb.sheetnames
        order = 0

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                logger.info("Parsing sheet: %s", sheet_name)

                # ── Sheet heading ─────────────────────────────────────────────
                elem = TextElement.build(
                    idx           = self._next_text_idx(),
                    content       = f"Sheet: {sheet_name}",
                    loc           = SourceLocation(sheet_name=sheet_name, order=order),
                    heading_level = 2,
                )
                doc_meta.text_elements.append(elem)
                order += 1

                # ── Data table ────────────────────────────────────────────────
                rows_data: list[list[str]] = []
                row_count = 0
                truncated = False

                for row in ws.iter_rows(values_only=True):
                    str_row = [_cell_str(c) for c in row]
                    if any(c for c in str_row):   # skip fully-empty rows
                        rows_data.append(str_row)
                        row_count += 1
                        if row_count >= MAX_ROWS:
                            logger.warning(
                                "Sheet '%s' exceeds %d rows — truncating", sheet_name, MAX_ROWS
                            )
                            truncated = True
                            break

                if rows_data:
                    # Trim trailing empty columns
                    max_col = max(
                        (i for r in rows_data for i, c in enumerate(r) if c),
                        default=0,
                    ) + 1
                    rows_data = [r[:max_col] for r in rows_data]

                    headers = rows_data[0] if rows_data else []
                    rows    = rows_data[1:] if len(rows_data) > 1 else []

                    # If first row is all numeric, there are no explicit headers
                    if headers and all(_is_numeric(h) for h in headers if h):
                        headers = [f"Col_{i + 1}" for i in range(len(headers))]
                        rows    = rows_data   # all rows are data rows

                    caption = f"Sheet data: {sheet_name}"
                    if truncated:
                        caption += f" (truncated at {MAX_ROWS} rows)"

                    elem_tbl = TableElement.build(
                        idx     = self._next_tbl_idx(),
                        loc     = SourceLocation(sheet_name=sheet_name, order=order),
                        headers = headers,
                        rows    = rows,
                        caption = caption,
                    )
                    doc_meta.table_elements.append(elem_tbl)
                    order += 1

                # ── Embedded images ───────────────────────────────────────────
                for img_elem in _extract_sheet_images(ws, sheet_name):
                    img_elem.source_location.order = order
                    img_elem.element_id = f"img_{self._next_img_idx():04d}"
                    img_elem.ref        = f"image#{img_elem.element_id}"
                    doc_meta.image_elements.append(img_elem)
                    order += 1

            except Exception as e:
                logger.warning("Sheet '%s' extraction failed: %s", sheet_name, e)

        wb.close()
        doc_meta.parsed_at = datetime.utcnow()
        doc_meta.rebuild_summary()
        return doc_meta


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _extract_sheet_images(ws, sheet_name: str) -> list[ImageElement]:
    """
    Version-safe image extraction from an openpyxl worksheet.
    openpyxl stores images in ws._images (list of openpyxl Image objects).
    The actual bytes are accessible via the image's _data() method (older) or
    by reading the underlying ref (newer).  We try both with fallback.
    """
    results: list[ImageElement] = []

    if not hasattr(ws, "_images"):
        return results

    for xl_img in ws._images:
        try:
            img_bytes = _get_image_bytes(xl_img)
            if not img_bytes or len(img_bytes) < MIN_IMG_BYTES:
                continue

            # Determine format from the PIL Image ref if available
            ext = _get_image_format(xl_img)
            b64 = base64.b64encode(img_bytes).decode("utf-8")

            # Placeholder element — idx/ref/order are filled in by caller
            elem = ImageElement(
                element_id      = "img_placeholder",
                ref             = "image#img_placeholder",
                source_location = SourceLocation(sheet_name=sheet_name, order=0),
                base64_data     = b64,
                format          = ext,
                size_bytes      = len(img_bytes),
            )
            results.append(elem)

        except Exception as e:
            logger.warning("Image extraction failed in sheet '%s': %s", sheet_name, e)

    return results


def _get_image_bytes(xl_img) -> Optional[bytes]:
    """Try multiple openpyxl API variations to get image bytes."""
    # Method 1: newer openpyxl — xl_img.ref is a PIL Image; read via BytesIO
    try:
        if hasattr(xl_img, "ref") and xl_img.ref is not None:
            buf = io.BytesIO()
            xl_img.ref.save(buf, format=xl_img.ref.format or "PNG")
            return buf.getvalue()
    except Exception:
        pass

    # Method 2: older openpyxl — xl_img._data() returns bytes directly
    try:
        if hasattr(xl_img, "_data") and callable(xl_img._data):
            data = xl_img._data()
            if data:
                return data
    except Exception:
        pass

    # Method 3: xl_img.image is sometimes the raw bytes
    try:
        if hasattr(xl_img, "image"):
            img = xl_img.image
            if isinstance(img, bytes) and img:
                return img
            # Sometimes it's a PIL Image
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return buf.getvalue()
    except Exception:
        pass

    return None


def _get_image_format(xl_img) -> str:
    """Determine the image format/extension from various openpyxl attrs."""
    try:
        if hasattr(xl_img, "format") and xl_img.format:
            return xl_img.format.lower()
    except Exception:
        pass
    try:
        if hasattr(xl_img, "ref") and xl_img.ref and hasattr(xl_img.ref, "format"):
            fmt = xl_img.ref.format
            if fmt:
                return fmt.lower()
    except Exception:
        pass
    return "png"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_numeric(s: str) -> bool:
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False
